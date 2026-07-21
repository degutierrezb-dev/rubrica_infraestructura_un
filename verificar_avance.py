#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verificar_avance.py
Cruza las evaluaciones recibidas en la carpeta de OneDrive con las asignaciones oficiales de Excel v3.0,
generando un reporte resumido en Markdown y un Dashboard interactivo en HTML.
"""

import os
import re
import json
import csv
import openpyxl
import unicodedata
from datetime import datetime

# Mapeos y Configuración
EXCEL_PATH = 'Detalle de espacios físicos para rúbrica v3.0.xlsx'
ONEDRIVE_FOLDER = r"C:\Users\degutierrez\OneDrive - Universidad del Norte\Migrado\Daniel Gutiérrez\Planta física\Evaluaciones recibidas"
SUMMARY_OUTPUT_PATH = r"C:\Users\degutierrez\.gemini\antigravity\brain\5c2c2dfa-09e3-440b-a02b-868aec769669\avance_real_summary.md"
HTML_OUTPUT_PATH = 'avance_real.html'

CAT_MAP_EXCEL_TO_ID = {
    "Salón de clases": "salon",
    "Laboratorios": "laboratorio",
    "Salas de Computadores": "sala_pc",
    "Fachadas de Edificios": "fachada",
    "Canchas Deportivas": "cancha",
    "Entradas y Salidas": "acceso",
    "Parqueaderos": "parqueadero",
    "Baños": "bano",
    "Cafeterías / Zonas de Alimentación": "cafeteria",
    "Bibliotecas / Salas de Estudio": "biblioteca",
    "Jardines y Zonas Verdes": "jardin",
    "Auditorios/Salas de Eventos": "auditorio",
}

CAT_MAP_ID_TO_NAME = {
    "salon": "Salones de Clases",
    "laboratorio": "Laboratorios",
    "sala_pc": "Salas de Computadores",
    "fachada": "Fachadas de Edificios",
    "cancha": "Canchas Deportivas",
    "acceso": "Entradas y Salidas",
    "parqueadero": "Parqueaderos",
    "bano": "Baños",
    "cafeteria": "Cafeterías / Zonas",
    "biblioteca": "Bibliotecas",
    "jardin": "Jardines y Zonas Verdes",
    "auditorio": "Auditorios / Salas de Eventos",
    "pasillo": "Pasillos/Escaleras/Áreas Comunes"
}

def normalize_text(text):
    if not text:
        return ""
    text = str(text)
    # Remove accents, lowercase, keep alphanumeric and spaces
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = text.lower()
    text = re.sub(r'[^a-z0-9]', ' ', text)
    return " ".join(text.split())

def clean_csv_header(h):
    # Normalize header names for matching
    norm = normalize_text(h)
    if 'codigo' in norm:
        return 'codigo'
    if 'nombre' in norm:
        return 'nombre'
    if 'categoria' in norm:
        return 'categoria'
    if 'edificio' in norm:
        return 'edificio'
    if 'piso' in norm:
        return 'piso'
    if 'fecha' in norm:
        return 'fecha'
    if 'hora' in norm:
        return 'hora'
    if 'evaluador' in norm:
        return 'evaluador'
    if 'cargo' in norm:
        return 'cargo'
    if 'promedio' in norm:
        return 'promedio'
    return norm

def main():
    print("=== Iniciando Verificación de Avance de Rúbrica ===")
    
    # 1. Cargar Asignaciones de Excel
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: No se encontró el archivo Excel de asignaciones en: {EXCEL_PATH}")
        return
        
    print(f"Cargando asignaciones desde {EXCEL_PATH}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['Asignaciones']
    except Exception as e:
        print(f"ERROR al cargar Excel: {e}")
        return

    assignments = []
    evaluadores_db = {}  # { normalized_name: official_name }
    evaluador_assignments = {} # { official_name: [list of assignments] }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        cat_excel, edificio, codigo, piso, nombre, _area, evaluador = row[:7]
        if not cat_excel or not evaluador:
            continue
            
        ev_name = str(evaluador).strip()
        ev_norm = normalize_text(ev_name)
        evaluadores_db[ev_norm] = ev_name
        
        if ev_name not in evaluador_assignments:
            evaluador_assignments[ev_name] = []
            
        cat_id = CAT_MAP_EXCEL_TO_ID.get(str(cat_excel).strip(), "otro")
        
        assign_item = {
            'cat': cat_id,
            'cat_excel': str(cat_excel).strip(),
            'ed': str(edificio).strip() if edificio else "",
            'cod': str(codigo).strip() if codigo else "",
            'piso': str(piso).strip() if piso is not None else "",
            'nom': str(nombre).strip() if nombre else "",
            'ev': ev_name
        }
        assignments.append(assign_item)
        evaluador_assignments[ev_name].append(assign_item)

    print(f"Se cargaron {len(assignments)} asignaciones oficiales de {len(evaluador_assignments)} evaluadores.")

    # 2. Cargar Evaluaciones desde OneDrive
    if not os.path.exists(ONEDRIVE_FOLDER):
        print(f"ERROR: No se encontró la carpeta de OneDrive: {ONEDRIVE_FOLDER}")
        return
        
    print(f"Escaneando archivos en {ONEDRIVE_FOLDER}...")
    evaluations = []
    
    # Track which file had how many evaluations
    file_stats = {}

    for filename in os.listdir(ONEDRIVE_FOLDER):
        path = os.path.join(ONEDRIVE_FOLDER, filename)
        if filename.endswith('.json'):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    evals = data.get('evaluations', [])
                    if not evals and isinstance(data, list):
                        evals = data
                    
                    count = 0
                    for ev in evals:
                        ev_evaluador = ev.get('evaluador', '').strip()
                        ev_codigo = ev.get('codigo', '').strip()
                        if not ev_evaluador or not ev_codigo:
                            continue
                            
                        evaluations.append({
                            'evaluador': ev_evaluador,
                            'codigo': ev_codigo,
                            'nombre': ev.get('nombre', ''),
                            'categoria': ev.get('categoria', ''),
                            'categoriaName': ev.get('categoriaName', ''),
                            'edificio': ev.get('edificio', ''),
                            'piso': ev.get('piso', ''),
                            'fecha': ev.get('fecha', ''),
                            'hora': ev.get('hora', ''),
                            'promedio': ev.get('promedio', 0),
                            'file': filename
                        })
                        count += 1
                    file_stats[filename] = count
            except Exception as e:
                print(f"ADVERTENCIA: Error al procesar JSON {filename}: {e}")
                
        elif filename.endswith('.csv'):
            try:
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = next(reader, None)
                    if headers:
                        clean_headers = [clean_csv_header(h) for h in headers]
                        count = 0
                        for row in reader:
                            if not row:
                                continue
                            row_dict = {}
                            for h, val in zip(clean_headers, row):
                                row_dict[h] = val
                            
                            ev_evaluador = row_dict.get('evaluador', '').strip()
                            ev_codigo = row_dict.get('codigo', '').strip()
                            if not ev_evaluador or not ev_codigo:
                                continue
                            
                            # Parse promedio
                            prom = 0
                            try:
                                prom = float(row_dict.get('promedio', 0))
                            except:
                                pass
                                
                            evaluations.append({
                                'evaluador': ev_evaluador,
                                'codigo': ev_codigo,
                                'nombre': row_dict.get('nombre', ''),
                                'categoria': row_dict.get('categoria', ''),
                                'categoriaName': row_dict.get('categoria', ''),
                                'edificio': row_dict.get('edificio', ''),
                                'piso': row_dict.get('piso', ''),
                                'fecha': row_dict.get('fecha', ''),
                                'hora': row_dict.get('hora', ''),
                                'promedio': prom,
                                'file': filename
                            })
                            count += 1
                        file_stats[filename] = count
            except Exception as e:
                print(f"ADVERTENCIA: Error al procesar CSV {filename}: {e}")

    print(f"Se cargaron {len(evaluations)} registros de evaluación en total.")

    # 3. Cruzar Datos y Calcular Avances
    # Create lookup map of evaluated spaces: { (normalized_evaluator_name, space_code): evaluation_item }
    evaluated_map = {}
    for ev in evaluations:
        key = (normalize_text(ev['evaluador']), ev['codigo'])
        evaluated_map[key] = ev

    # We will compute results for each of the 39 official evaluators
    evaluators_stats = []
    
    total_assigned_spaces = len(assignments)
    total_completed_assigned = 0
    total_extra_evaluated = 0

    # Sort evaluators alphabetically
    for ev_name in sorted(evaluador_assignments.keys()):
        ev_norm = normalize_text(ev_name)
        ev_assigns = evaluador_assignments[ev_name]
        
        assigned_count = len(ev_assigns)
        done_list = []
        pending_list = []
        extra_list = []
        
        # Check assigned spaces
        for assign in ev_assigns:
            code = assign['cod']
            lookup_key = (ev_norm, code)
            
            if lookup_key in evaluated_map:
                ev_data = evaluated_map[lookup_key]
                done_list.append({
                    'cod': code,
                    'nom': assign['nom'],
                    'cat': assign['cat_excel'],
                    'ed': assign['ed'],
                    'piso': assign['piso'],
                    'fecha': ev_data['fecha'],
                    'promedio': ev_data['promedio'],
                    'file': ev_data['file']
                })
                total_completed_assigned += 1
            else:
                pending_list.append(assign)
                
        # Find extra evaluations by this evaluator (spaces they evaluated but weren't assigned to them)
        assigned_codes_set = set(a['cod'] for a in ev_assigns)
        for key, ev_data in evaluated_map.items():
            ev_norm_key, code = key
            if ev_norm_key == ev_norm and code not in assigned_codes_set:
                extra_list.append({
                    'cod': code,
                    'nom': ev_data['nombre'],
                    'cat': ev_data['categoriaName'] or CAT_MAP_ID_TO_NAME.get(ev_data['categoria'], ev_data['categoria']),
                    'ed': ev_data['edificio'],
                    'piso': ev_data['piso'],
                    'fecha': ev_data['fecha'],
                    'promedio': ev_data['promedio'],
                    'file': ev_data['file']
                })
                total_extra_evaluated += 1

        progress_pct = (len(done_list) / assigned_count * 100) if assigned_count > 0 else 0
        
        evaluators_stats.append({
            'name': ev_name,
            'assigned_count': assigned_count,
            'evaluated_count': len(done_list),
            'pending_count': len(pending_list),
            'extra_count': len(extra_list),
            'progress_pct': progress_pct,
            'done_list': done_list,
            'pending_list': pending_list,
            'extra_list': extra_list
        })

    # Global stats
    global_progress_pct = (total_completed_assigned / total_assigned_spaces * 100) if total_assigned_spaces > 0 else 0

    print(f"Avance Global: {total_completed_assigned}/{total_assigned_spaces} ({global_progress_pct:.1f}%)")
    print(f"Espacios extras/ad-hoc evaluados: {total_extra_evaluated}")

    # 4. Generar reporte avance_real_summary.md
    generate_markdown_report(total_assigned_spaces, total_completed_assigned, total_extra_evaluated, global_progress_pct, evaluators_stats, file_stats)

    # 5. Generar dashboard interactivo avance_real.html
    generate_html_dashboard(total_assigned_spaces, total_completed_assigned, total_extra_evaluated, global_progress_pct, evaluators_stats)

    print("\n=== Verificación Completada Exitosamente ===")
    print(f"Reporte Markdown escrito en: {SUMMARY_OUTPUT_PATH}")
    print(f"Dashboard interactivo escrito en: {HTML_OUTPUT_PATH}")

def generate_markdown_report(total_assigned, total_done, total_extra, global_pct, ev_stats, file_stats):
    lines = []
    lines.append("# Reporte de Monitoreo de Avance en Tiempo Real")
    lines.append(f"\n*Generado automáticamente el {datetime.now().strftime('%d/%m/%Y a las %I:%M %p')}*")
    lines.append("\n---")
    
    lines.append("\n## 📊 Resumen Ejecutivo")
    lines.append(f"- **Total de Espacios Asignados:** {total_assigned}")
    lines.append(f"- **Evaluaciones de Asignación Completadas:** {total_done}")
    lines.append(f"- **Avance de Asignaciones:** `{global_pct:.1f}%`")
    lines.append(f"- **Espacios Extras Evaluados (ej. Pasillos/Otras áreas):** {total_extra}")
    lines.append(f"- **Total de Registros de Evaluación Consolidados:** {total_done + total_extra}")

    # File lists
    lines.append("\n### 📂 Archivos JSON/CSV Procesados en OneDrive")
    lines.append("| Archivo | Registros Encontrados |")
    lines.append("| :--- | :---: |")
    for fname, count in sorted(file_stats.items()):
        lines.append(f"| {fname} | {count} |")

    # Evaluators summary table
    lines.append("\n## 👥 Resumen por Evaluador/a")
    lines.append("| Evaluador/a | Asignados | Evaluados | Pendientes | Extras | Avance |")
    lines.append("| :--- | :---: | :---: | :---: | :---: | :---: |")
    for ev in ev_stats:
        status_emoji = "🟢" if ev['progress_pct'] >= 100 else ("🟡" if ev['progress_pct'] > 0 else "🔴")
        lines.append(f"| {ev['name']} | {ev['assigned_count']} | {ev['evaluated_count']} | {ev['pending_count']} | {ev['extra_count']} | {status_emoji} `{ev['progress_pct']:.1f}%` |")

    lines.append("\n---")
    lines.append("\n> [!NOTE]\n> Si un evaluador tiene espacios pendientes, consulte el archivo local `avance_real.html` para ver el listado detallado de cuáles códigos/espacios específicos le hacen falta.")

    with open(SUMMARY_OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))

def generate_html_dashboard(total_assigned, total_done, total_extra, global_pct, ev_stats):
    stats_json_str = json.dumps(ev_stats, ensure_ascii=False)

    html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <title>Dashboard de Monitoreo de Avance — UniNorte</title>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg: #0a0e1a; --bg2: #111827; --bg-card: rgba(17,24,39,0.7);
      --glass: rgba(255,255,255,0.04); --border: rgba(255,255,255,0.08);
      --text: #f1f5f9; --text2: #94a3b8; --muted: #64748b;
      --accent: #0ea5e9; --accent2: #06b6d4;
      --green: #22c55e; --blue: #3b82f6; --yellow: #eab308;
      --orange: #f97316; --red: #ef4444;
      --radius: 12px;
    }}
    
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    
    body {{
      font-family: 'Inter', sans-serif;
      background: var(--bg);
      color: var(--text);
      line-height: 1.5;
      padding-bottom: 60px;
    }}

    header {{
      background: rgba(10,14,26,0.9);
      backdrop-filter: blur(20px);
      border-bottom: 1px solid var(--border);
      padding: 20px 24px;
      position: sticky;
      top: 0;
      z-index: 50;
    }}
    .header-inner {{
      max-width: 1200px;
      margin: 0 auto;
      display: flex;
      justify-content: space-between;
      align-items: center;
      flex-wrap: wrap;
      gap: 16px;
    }}
    header h1 {{
      font-size: 1.25rem;
      font-weight: 700;
      background: linear-gradient(135deg, var(--accent), var(--accent2));
      -webkit-background-clip: text;
      -webkit-text-fill-color: transparent;
      background-clip: text;
    }}
    header p {{ font-size: 0.75rem; color: var(--muted); margin-top: 2px; }}

    main {{
      max-width: 1200px;
      margin: 0 auto;
      padding: 24px;
    }}

    /* Stats Grid */
    .stats-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 16px;
      margin-bottom: 24px;
    }}
    .stat-card {{
      background: var(--bg-card);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      padding: 20px;
      backdrop-filter: blur(12px);
      position: relative;
      overflow: hidden;
    }}
    .stat-label {{
      font-size: 0.72rem;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.06em;
      font-weight: 600;
    }}
    .stat-value {{
      font-size: 2rem;
      font-weight: 800;
      margin-top: 6px;
      line-height: 1;
    }}
    .stat-sub {{
      font-size: 0.75rem;
      color: var(--text2);
      margin-top: 6px;
    }}
    .stat-accent {{
      position: absolute;
      top: 0;
      right: 0;
      width: 60px;
      height: 60px;
      border-radius: 0 0 0 60px;
      opacity: 0.1;
    }}

    /* Search & Controls */
    .controls-row {{
      display: flex;
      gap: 16px;
      margin-bottom: 20px;
      align-items: center;
      flex-wrap: wrap;
    }}
    .search-input {{
      flex: 1;
      min-width: 280px;
      background: var(--bg-card);
      border: 1px solid var(--border);
      border-radius: 8px;
      color: var(--text);
      padding: 10px 16px;
      font-family: inherit;
      font-size: 0.88rem;
    }}
    .search-input:focus {{
      outline: none;
      border-color: var(--accent);
    }}
    .filter-btn {{
      padding: 10px 16px;
      border-radius: 8px;
      font-family: inherit;
      font-size: 0.82rem;
      font-weight: 600;
      cursor: pointer;
      border: 1px solid var(--border);
      background: var(--glass);
      color: var(--text2);
      transition: all 0.2s;
    }}
    .filter-btn.active {{
      background: var(--accent);
      border-color: var(--accent);
      color: white;
    }}
    .sort-select {{
      padding: 10px 16px;
      border-radius: 8px;
      font-family: inherit;
      font-size: 0.82rem;
      font-weight: 600;
      cursor: pointer;
      border: 1px solid var(--border);
      background: var(--bg-card);
      color: var(--text2);
      outline: none;
      transition: all 0.2s;
    }}
    .sort-select:focus {{
      border-color: var(--accent);
    }}
    .sort-select option {{
      background: var(--bg2);
      color: var(--text);
    }}

    /* Evaluator Row */
    .ev-card {{
      background: var(--bg-card);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      margin-bottom: 12px;
      overflow: hidden;
      transition: border-color 0.2s;
    }}
    .ev-card:hover {{
      border-color: rgba(255,255,255,0.15);
    }}
    .ev-header {{
      padding: 16px 20px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      cursor: pointer;
      user-select: none;
      flex-wrap: wrap;
      gap: 16px;
    }}
    .ev-info {{
      display: flex;
      align-items: center;
      gap: 12px;
    }}
    .ev-status {{
      font-size: 1.25rem;
    }}
    .ev-name {{
      font-weight: 600;
      font-size: 0.95rem;
    }}
    .ev-stats-brief {{
      font-size: 0.78rem;
      color: var(--text2);
    }}
    .progress-container {{
      display: flex;
      align-items: center;
      gap: 12px;
      min-width: 250px;
    }}
    .progress-bar-bg {{
      flex: 1;
      height: 8px;
      background: rgba(255,255,255,0.06);
      border-radius: 4px;
      overflow: hidden;
    }}
    .progress-bar-fill {{
      height: 100%;
      border-radius: 4px;
      transition: width 0.5s;
    }}
    .progress-text {{
      font-size: 0.78rem;
      font-weight: 700;
      width: 45px;
      text-align: right;
    }}

    /* Expandable details */
    .ev-details {{
      display: none;
      border-top: 1px solid var(--border);
      background: rgba(10,14,26,0.5);
      padding: 20px;
    }}
    .ev-card.open .ev-details {{
      display: block;
    }}

    /* Tabs */
    .tabs {{
      display: flex;
      gap: 8px;
      border-bottom: 1px solid var(--border);
      margin-bottom: 16px;
    }}
    .tab {{
      padding: 8px 16px;
      font-size: 0.8rem;
      font-weight: 600;
      cursor: pointer;
      color: var(--muted);
      border-bottom: 2px solid transparent;
      transition: all 0.2s;
    }}
    .tab:hover {{
      color: var(--text);
    }}
    .tab.active {{
      color: var(--accent);
      border-bottom-color: var(--accent);
    }}
    .tab-content {{
      display: none;
    }}
    .tab-content.active {{
      display: block;
    }}

    /* Tables */
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 0.78rem;
    }}
    thead th {{
      text-align: left;
      padding: 8px 12px;
      color: var(--muted);
      font-size: 0.68rem;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      border-bottom: 1px solid var(--border);
    }}
    tbody tr {{
      border-bottom: 1px solid rgba(255,255,255,0.02);
    }}
    tbody tr:last-child {{
      border-bottom: none;
    }}
    td {{
      padding: 8px 12px;
      vertical-align: middle;
    }}
    .score-badge {{
      display: inline-block;
      padding: 2px 6px;
      border-radius: 4px;
      font-weight: 700;
      font-size: 0.72rem;
    }}
    .file-badge {{
      font-size: 0.65rem;
      color: var(--muted);
      background: var(--glass);
      padding: 2px 6px;
      border-radius: 4px;
      border: 1px solid var(--border);
    }}
    
    .empty-tab {{
      text-align: center;
      padding: 30px 20px;
      color: var(--muted);
      font-size: 0.82rem;
    }}
  </style>
</head>
<body>

  <header>
    <div class="header-inner">
      <div>
        <h1>📊 Monitoreo de Avance en Tiempo Real</h1>
        <p>Rúbrica de Infraestructura — Universidad del Norte</p>
      </div>
      <div>
        <p style="text-align: right; color: var(--text2);">Actualizado: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}</p>
      </div>
    </div>
  </header>

  <main>
    <!-- Stats Cards -->
    <div class="stats-grid">
      <div class="stat-card">
        <div class="stat-label">Espacios Asignados</div>
        <div class="stat-value">{total_assigned}</div>
        <div class="stat-sub">Registrados en Excel v3.0</div>
        <div class="stat-accent" style="background:var(--blue);"></div>
      </div>
      <div class="stat-card">
        <div class="stat-label">Evaluados de Asignación</div>
        <div class="stat-value" style="color: var(--green);">{total_done}</div>
        <div class="stat-sub">Avance general: {global_pct:.1f}%</div>
        <div class="stat-accent" style="background:var(--green);"></div>
      </div>
      <div class="stat-card">
        <div class="stat-label">Avance Global</div>
        <div class="stat-value" style="color: var(--accent);">{global_pct:.1f}%</div>
        <div class="progress-bar-bg" style="margin-top:10px; height:6px;">
          <div class="progress-bar-fill" style="width: {global_pct}%; background: var(--accent);"></div>
        </div>
        <div class="stat-accent" style="background:var(--accent);"></div>
      </div>
      <div class="stat-card">
        <div class="stat-label">Espacios Extras Evaluados</div>
        <div class="stat-value" style="color: var(--yellow);">{total_extra}</div>
        <div class="stat-sub">Pasillos y ad-hoc fuera de planilla</div>
        <div class="stat-accent" style="background:var(--yellow);"></div>
      </div>
    </div>

    <!-- Controls -->
    <div class="controls-row">
      <input type="text" class="search-input" id="searchInput" placeholder="🔍 Buscar evaluador/a..." oninput="filterEvaluators()">
      <button class="filter-btn active" id="btnFilterAll" onclick="setFilter('all')">Todos</button>
      <button class="filter-btn" id="btnFilterPending" onclick="setFilter('pending')">Con pendientes</button>
      <button class="filter-btn" id="btnFilterComplete" onclick="setFilter('complete')">Completados (100%)</button>
      <select class="sort-select" id="sortSelect" onchange="handleSortChange()">
        <option value="name_asc">Ordenar: Nombre (A-Z)</option>
        <option value="name_desc">Ordenar: Nombre (Z-A)</option>
        <option value="progress_desc">Ordenar: Avance (Mayor a Menor)</option>
        <option value="progress_asc">Ordenar: Avance (Menor a Mayor)</option>
      </select>
    </div>

    <!-- Evaluators List -->
    <div id="evaluatorsContainer"></div>
  </main>

  <script>
    const DATA = {stats_json_str};

    let activeFilter = 'all'; // all, pending, complete
    let searchQuery = '';
    let currentSort = 'name_asc';

    function getProgressColor(pct) {{
      if (pct >= 100) return 'var(--green)';
      if (pct >= 50) return 'var(--yellow)';
      return 'var(--red)';
    }}

    function getProgressEmoji(pct) {{
      if (pct >= 100) return '🟢';
      if (pct > 0) return '🟡';
      return '🔴';
    }}

    function getScoreClass(score) {{
      if (score >= 4.5) return {{ label: "Excelente", color: "#22c55e" }};
      if (score >= 3.5) return {{ label: "Bueno", color: "#3b82f6" }};
      if (score >= 2.5) return {{ label: "Aceptable", color: "#eab308" }};
      if (score >= 1.5) return {{ label: "Deficiente", color: "#f97316" }};
      return {{ label: "Crítico", color: "#ef4444" }};
    }}

    function toggleCard(idx) {{
      const card = document.getElementById(`ev-card-${{idx}}`);
      if (card) card.classList.toggle('open');
    }}

    function switchTab(evIdx, tabName) {{
      document.querySelectorAll(`#ev-card-${{evIdx}} .tab`).forEach(t => t.classList.remove('active'));
      document.querySelectorAll(`#ev-card-${{evIdx}} .tab-content`).forEach(c => c.classList.remove('active'));

      document.querySelector(`#ev-card-${{evIdx}} .tab[data-tab="${{tabName}}"]`).classList.add('active');
      document.getElementById(`tab-content-${{evIdx}}-${{tabName}}`).classList.add('active');
      
      event.stopPropagation();
    }}

    function renderEvaluators() {{
      const container = document.getElementById('evaluatorsContainer');
      container.innerHTML = '';

      const filteredData = DATA.filter(ev => {{
        const matchQuery = ev.name.toLowerCase().includes(searchQuery.toLowerCase());
        
        let matchFilter = true;
        if (activeFilter === 'pending') {{
          matchFilter = ev.evaluated_count < ev.assigned_count;
        }} else if (activeFilter === 'complete') {{
          matchFilter = ev.evaluated_count >= ev.assigned_count;
        }}
        
        return matchQuery && matchFilter;
      }});

      filteredData.sort((a, b) => {{
        if (currentSort === 'name_asc') {{
          return a.name.localeCompare(b.name);
        }} else if (currentSort === 'name_desc') {{
          return b.name.localeCompare(a.name);
        }} else if (currentSort === 'progress_desc') {{
          if (b.progress_pct !== a.progress_pct) {{
            return b.progress_pct - a.progress_pct;
          }}
          return a.name.localeCompare(b.name);
        }} else if (currentSort === 'progress_asc') {{
          if (a.progress_pct !== b.progress_pct) {{
            return a.progress_pct - b.progress_pct;
          }}
          return a.name.localeCompare(b.name);
        }}
        return 0;
      }});

      if (filteredData.length === 0) {{
        container.innerHTML = `<div style="text-align:center;padding:40px;color:var(--muted);">No se encontraron evaluadores con los criterios seleccionados.</div>`;
        return;
      }}

      filteredData.forEach((ev, i) => {{
        const color = getProgressColor(ev.progress_pct);
        const emoji = getProgressEmoji(ev.progress_pct);
        
        let pendingHtml = '';
        if (ev.pending_list.length > 0) {{
          pendingHtml = `
            <table>
              <thead>
                <tr>
                  <th>Código</th>
                  <th>Nombre</th>
                  <th>Categoría</th>
                  <th>Edificio</th>
                  <th>Piso</th>
                </tr>
              </thead>
              <tbody>
                ${{ev.pending_list.map(p => `
                  <tr>
                    <td><code>${{p.cod}}</code></td>
                    <td><strong>${{p.nom}}</strong></td>
                    <td>${{p.cat_excel}}</td>
                    <td>${{p.ed}}</td>
                    <td>${{p.piso || '—'}}</td>
                  </tr>
                `).join('')}}
              </tbody>
            </table>
          `;
        }} else {{
          pendingHtml = `<div class="empty-tab">🎉 ¡Todos los espacios asignados han sido evaluados!</div>`;
        }}

        let doneHtml = '';
        if (ev.done_list.length > 0) {{
          doneHtml = `
            <table>
              <thead>
                <tr>
                  <th>Código</th>
                  <th>Nombre</th>
                  <th>Categoría</th>
                  <th>Edificio</th>
                  <th>Fecha</th>
                  <th style="text-align:center;">Puntaje</th>
                  <th>Archivo</th>
                </tr>
              </thead>
              <tbody>
                ${{ev.done_list.map(d => {{
                  const cls = getScoreClass(d.promedio);
                  return `
                    <tr>
                      <td><code>${{d.cod}}</code></td>
                      <td><strong>${{d.nom}}</strong></td>
                      <td>${{d.cat}}</td>
                      <td>${{d.ed}}</td>
                      <td>${{d.fecha}}</td>
                      <td style="text-align:center;">
                        <span class="score-badge" style="background:${{cls.color}}22; color:${{cls.color}};">${{d.promedio.toFixed(2)}}</span>
                      </td>
                      <td><span class="file-badge">${{d.file}}</span></td>
                    </tr>
                  `;
                }}).join('')}}
              </tbody>
            </table>
          `;
        }} else {{
          doneHtml = `<div class="empty-tab">Ningún espacio asignado ha sido evaluado aún.</div>`;
        }}

        let extraHtml = '';
        if (ev.extra_list.length > 0) {{
          extraHtml = `
            <table>
              <thead>
                <tr>
                  <th>Código</th>
                  <th>Nombre</th>
                  <th>Categoría</th>
                  <th>Edificio</th>
                  <th>Fecha</th>
                  <th style="text-align:center;">Puntaje</th>
                  <th>Archivo</th>
                </tr>
              </thead>
              <tbody>
                ${{ev.extra_list.map(ex => {{
                  const cls = getScoreClass(ex.promedio);
                  return `
                    <tr>
                      <td><code>${{ex.cod}}</code></td>
                      <td><strong>${{ex.nom}}</strong></td>
                      <td>${{ex.cat}}</td>
                      <td>${{ex.ed}}</td>
                      <td>${{ex.fecha}}</td>
                      <td style="text-align:center;">
                        <span class="score-badge" style="background:${{cls.color}}22; color:${{cls.color}};">${{ex.promedio.toFixed(2)}}</span>
                      </td>
                      <td><span class="file-badge">${{ex.file}}</span></td>
                    </tr>
                  `;
                }}).join('')}}
              </tbody>
            </table>
          `;
        }} else {{
          extraHtml = `<div class="empty-tab">No hay evaluaciones extras (ej. Pasillos o espacios de terceros) registradas para este evaluador.</div>`;
        }}

        const card = document.createElement('div');
        card.className = 'ev-card';
        card.id = `ev-card-${{i}}`;
        card.innerHTML = `
          <div class="ev-header" onclick="toggleCard(${{i}})">
            <div class="ev-info">
              <span class="ev-status">${{emoji}}</span>
              <div>
                <div class="ev-name">${{ev.name}}</div>
                <div class="ev-stats-brief">
                  Asignados: ${{ev.assigned_count}} · Evaluados: ${{ev.evaluated_count}} · Pendientes: ${{ev.pending_count}} · Extras: ${{ev.extra_count}}
                </div>
              </div>
            </div>
            <div class="progress-container">
              <div class="progress-bar-bg">
                <div class="progress-bar-fill" style="width: ${{ev.progress_pct}}%; background: ${{color}};"></div>
              </div>
              <div class="progress-text" style="color: ${{color}};">${{ev.progress_pct.toFixed(0)}}%</div>
            </div>
          </div>
          
          <div class="ev-details">
            <div class="tabs">
              <div class="tab active" data-tab="pending" onclick="switchTab(${{i}}, 'pending')">Pendientes (${{ev.pending_count}})</div>
              <div class="tab" data-tab="done" onclick="switchTab(${{i}}, 'done')">Evaluados de Asignación (${{ev.evaluated_count}})</div>
              <div class="tab" data-tab="extra" onclick="switchTab(${{i}}, 'extra')">Evaluaciones Extras (${{ev.extra_count}})</div>
            </div>
            
            <div class="tab-content active" id="tab-content-${{i}}-pending">
              ${{pendingHtml}}
            </div>
            <div class="tab-content" id="tab-content-${{i}}-done">
              ${{doneHtml}}
            </div>
            <div class="tab-content" id="tab-content-${{i}}-extra">
              ${{extraHtml}}
            </div>
          </div>
        `;
        container.appendChild(card);
      }});
    }}

    function filterEvaluators() {{
      searchQuery = document.getElementById('searchInput').value;
      renderEvaluators();
    }}

    function setFilter(filterType) {{
      activeFilter = filterType;
      document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
      
      if (filterType === 'all') document.getElementById('btnFilterAll').classList.add('active');
      else if (filterType === 'pending') document.getElementById('btnFilterPending').classList.add('active');
      else if (filterType === 'complete') document.getElementById('btnFilterComplete').classList.add('active');
      
      renderEvaluators();
    }}

    function handleSortChange() {{
      currentSort = document.getElementById('sortSelect').value;
      renderEvaluators();
    }}

    renderEvaluators();
  </script>
</body>
</html>
"""

    with open(HTML_OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(html_content)

if __name__ == '__main__':
    main()
