#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera las constantes ESPACIOS y EVALUADORES como código JavaScript
a partir de la hoja 'Asignaciones' del Excel v3.0.
"""
import openpyxl

# Mapeo de categoría Excel -> id de la rúbrica
CAT_MAP = {
    "Salón de clases": "salon",
    "Laboratorios": "laboratorio",
    "Salas de Computadores": "sala_pc",
    "Fachadas de Edificios": "fachada",
    "Canchas Deportivas": "cancha",
    "Entradas y Salidas": "acceso",
    "Parqueaderos": "parqueadero",
    "Baños": "bano",
    "Cafeterías / Zonas de Alimentación": "cafeteria",
    "Bibliotecas / Salas de Estudio": "biblioteca",
    "Jardines y Zonas Verdes": "jardin",
    "Auditorios/Salas de Eventos": "auditorio",
}

EXCEL_PATH = r'Detalle de espacios físicos para rúbrica v3.0.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Asignaciones']

lines = []
lines.append("// Espacios físicos extraídos del archivo Excel v3.0 (hoja Asignaciones)")
lines.append("// Generado automáticamente — no editar manualmente")
lines.append("const ESPACIOS = [")

evaluadores_set = set()
count = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    cat_excel, edificio, codigo, piso, nombre, _area, evaluador = row[:7]
    if not cat_excel:
        continue
    
    cat_id = CAT_MAP.get(str(cat_excel).strip(), "")
    if not cat_id:
        print(f"ADVERTENCIA: Categoría no mapeada: '{cat_excel}'")
        continue
    
    ed = str(edificio).strip() if edificio else ""
    cod = str(codigo).strip() if codigo else ""
    piso_str = str(piso).strip() if piso is not None else ""
    # Clean non-standard dashes
    if piso_str in ('—', '–', '-', 'N/A', 'n/a', 'None'):
        piso_str = ""
    nom = str(nombre).strip() if nombre else ""
    ev = str(evaluador).strip() if evaluador else ""
    
    # Track evaluadores
    if ev:
        evaluadores_set.add(ev)
    
    # Escape quotes in strings
    ed = ed.replace('"', '\\"')
    cod = cod.replace('"', '\\"')
    piso_str = piso_str.replace('"', '\\"')
    nom = nom.replace('"', '\\"')
    ev = ev.replace('"', '\\"')
    
    lines.append(f'  {{cat:"{cat_id}",ed:"{ed}",cod:"{cod}",piso:"{piso_str}",nom:"{nom}",ev:"{ev}"}},')
    count += 1

lines.append("];")
lines.append("")

# Generate EVALUADORES constant (sorted list of unique names)
evaluadores_sorted = sorted(evaluadores_set)
lines.append("// Lista de evaluadores extraída de la columna 'Evaluador/a'")
lines.append("const EVALUADORES = [")
for ev_name in evaluadores_sorted:
    ev_escaped = ev_name.replace('"', '\\"')
    lines.append(f'  "{ev_escaped}",')
lines.append("];")
lines.append("")

output = "\n".join(lines)

with open("espacios_generated.js", "w", encoding="utf-8") as f:
    f.write(output)

print(f"Generado espacios_generated.js ({count} espacios, {len(evaluadores_sorted)} evaluadores)")
